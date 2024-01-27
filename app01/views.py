import json
import random
import datetime
from django.shortcuts import render, HttpResponse, redirect
from app01.models import User, Department, Pretty, Admin, Order
from app01.utils.pagiation import Pagination
from app01.form.form import PrettyModelForm, PrettyEditModelForm, UserModelForm, AdminModelForm, AdminEditModelForm, \
    AdminResetModelForm, LoginForm, OrderModelForm, ExcelForm, RegisterForm, SendSmsForm, SmsLoginForm
from django.views.decorators.csrf import csrf_exempt
from app01.utils.code import check_code
from django.shortcuts import HttpResponse
from io import BytesIO
import json
from openpyxl import load_workbook
from django_redis import get_redis_connection


# Create your views here.
def index(request):
    # conn = get_redis_connection('default')  # default是连接池的名称
    # conn.set("name", "冰冷的希望")
    # name = conn.get('name').decode('utf-8')
    # print(name)
    return HttpResponse('首页')


# #################################   部门管理  # #################################
def department_list(request):
    """部门列表"""
    department_objs = Department.objects.all()
    return render(request, 'department_list.html', {'department_info': department_objs})


def department_add(request):
    """部门添加"""
    if request.method == "GET":
        return render(request, 'department_add.html')
    depart_name = request.POST.get('depart_name')
    # 保存到数据库
    Department.objects.create(name=depart_name)
    return redirect("/department/list/")


def department_multi(request):
    # 获取上传的 Excel 文件对象
    file_object = request.FILES.get('exc')
    print(request.FILES)
    form = ExcelForm(data=file_object.name, files=request.FILES)
    print(form.errors)
    print(form.is_valid())
    if form.is_valid():
        file_object = request.FILES.get('exc')
        print(file_object)
        if 'xlsx' not in file_object.name:
            return HttpResponse("文件格式有误")
        file_object = form.cleaned_data.get("exc")
        print(file_object)
        wb = load_workbook(file_object)
        sheet = wb.worksheets[0]
        # 循环获取每一行数据,并更新至数据库
        for row in sheet.iter_rows(min_row=2):
            exc_title = row[0].value
            # 如果表格中的数据在数据库中不存在,则进行创建
            if not Department.objects.filter(name=exc_title).exists():
                Department.objects.create(name=exc_title)
    file_object = request.FILES.get('exc')
    if 'xlsx' not in file_object.name:
        return HttpResponse("文件格式有误")
    # 打开 Excel 文件读取内容
    return redirect('/department/list/')


def department_delete(request):
    """部门删除"""

    nid = request.GET.get('nid')
    Department.objects.filter(id=nid).delete()
    # 重定向回部门列表
    return redirect("/department/list/")


def department_edit(request, nid):
    """部门编辑"""
    if request.method == "POST":
        # post请求将修改数据提交到数据库
        # 如果是POST请求,保存修改
        depart_name = request.POST.get('depart_name')
        Department.objects.filter(id=nid).update(name=depart_name)
        return redirect("/department/list/")
    # get请求，返回编辑页面
    department_obj = Department.objects.filter(id=nid).get()
    return render(request, 'department_edit.html', {'department_obj': department_obj})


# #################################   用户管理  # #################################
def user_list(request):
    # 获取所有用户列表
    user_data = User.objects.all()
    for user in user_data:
        print(user.get_gender_display())
    return render(request, "user_list.html", {"user_data": user_data})


def user_add(request):
    """用户添加"""
    if request.method == "GET":
        form = UserModelForm()
        return render(request, 'user_add.html', {'form': form})
        # 使用ModelForm进行验证POST请求提交数据
    form = UserModelForm(data=request.POST)
    if form.is_valid():
        # 直接保存至数据库
        form.save()
        return redirect("/user/list/")
    # 校验失败(在页面上显示错误信息)
    return render(request, "user_add.html", {"form": form})


def user_delete(request):
    """用户删除"""
    nid = request.POST.get('nid')
    User.objects.filter(id=nid).delete()
    return redirect("/user/list/")


def user_edit(request, nid):
    """编辑用户"""
    row_obj = User.objects.filter(id=nid).first()
    # GET请求
    if request.method == "GET":
        form = UserModelForm(instance=row_obj)
        return render(request, "user_edit.html", {"form": form})
    # POST请求
    form = UserModelForm(data=request.POST, instance=row_obj)
    if form.is_valid():
        form.save()
        return redirect("/user/list/")
    return render(request, "user_edit.html", {"form": form})


# #################################   靓号管理  # #################################
def pretty_list(request):
    """
    靓号列表
    :param request:
    :return:
    """

    # 获取所有用户列表
    data_dic = {}
    query = request.GET.get('query', "")
    if query:
        data_dic['mobile__contains'] = query
    queryset = Pretty.objects.filter(**data_dic).order_by("-level")
    # 分页操作
    page_obj = Pagination(request, queryset, 2)
    # 分页后的数据
    page_pretty_data = page_obj.page_data
    # 分页html页面
    page_html = page_obj.html()
    return render(request, "pretty_list.html",
                  {"pretty_data": page_pretty_data, 'query': query, "page_string": page_html})


def pretty_add(request):
    """
    靓号添加
    :param request:
    :return:
    """
    if request.method == "GET":
        form = PrettyModelForm()
        return render(request, "pretty_add.html", {"form": form})
        # 使用ModelForm进行验证POST请求提交数据
    form = PrettyModelForm(data=request.POST)
    if form.is_valid():
        # 直接保存至数据库
        form.save()
        return redirect("/pretty/list/")
    # 校验失败(在页面上显示错误信息)
    return render(request, "pretty_add.html", {"form": form})


def pretty_edit(request, nid):
    """
   靓号编辑
   :param request:
   :return:
   """
    """编辑用户"""
    row_obj = Pretty.objects.filter(id=nid).first()
    # GET请求
    if request.method == "GET":
        form = PrettyEditModelForm(instance=row_obj)
        return render(request, "pretty_edit.html", {"form": form})
    # POST请求
    form = PrettyEditModelForm(data=request.POST, instance=row_obj)
    if form.is_valid():
        form.save()
        return redirect("/pretty/list/")
    return render(request, "pretty_edit.html", {"form": form})


def pretty_delete(request, nid):
    """
   靓号删除
   :param request:
   :return:
   """
    Pretty.objects.filter(id=nid).delete()
    return redirect("/pretty/list/")


def admin_list(request):
    """
    管理员列表
    :param request:
    :return:
    """
    # 获取所有管理员列表
    data_dic = {}
    query = request.GET.get('query', "")
    if query:
        # 筛选出包含
        data_dic['name__contains'] = query
    queryset = Admin.objects.filter(**data_dic)
    # 分页操作
    page_obj = Pagination(request, queryset, 2)
    # 分页后的数据
    page_data = page_obj.page_data
    # 分页html页面
    page_html = page_obj.html()
    return render(request, "admin_list.html",
                  {"page_data": page_data, 'query': query, "page_string": page_html})


def admin_edit(request, nid):
    """
    管理员编辑页面
    :param request:
    :return:
    """
    # 判断 nid 是否存在
    row_object = Admin.objects.filter(id=nid).first()
    if request.method == "GET":
        form = AdminEditModelForm(instance=row_object)
        return render(request, "admin_edit.html", {"form": form})
    form = AdminEditModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/admin/list/')
    return render(request, "admin_edit.html", {"form": form})


def admin_add(request):
    """
    管理员添加页面
    :param request:
    :return:
    """
    if request.method == "GET":
        form = AdminModelForm()
        return render(request, "admin_add.html", {"form": form})
        # 使用ModelForm进行验证POST请求提交数据
    form = AdminModelForm(data=request.POST)
    if form.is_valid():
        # 直接保存至数据库
        form.save()
        return redirect("/admin/list/")
    # 校验失败(在页面上显示错误信息)
    return render(request, "admin_add.html", {"form": form})


def admin_delete(request, nid):
    Admin.objects.filter(id=nid).delete()
    return redirect("/admin/list/")


def admin_reset(request, nid):
    row_object = Admin.objects.filter(id=nid).first()
    title = "重置密码 - {}".format(row_object.name)
    if request.method == "GET":
        form = AdminResetModelForm(instance=row_object)
        return render(request, "admin_reset.html", {'form': form})
    form = AdminResetModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect("/admin/list")
    return render(request, "admin_reset.html", {'form': form, 'title': title})


def login(request):
    if request.method == "GET":
        form = LoginForm()
        return render(request, "login.html", {"form": form})
    form = LoginForm(data=request.POST)
    if form.is_valid():
        # 验证码的校验
        input_code = form.cleaned_data.pop("code")
        img_code = request.session.get("image_code", "")
        print("user_input_code={}, image_code={}".format(input_code, img_code))
        # 用户名和密码校验
        admin_object = Admin.objects.filter(**form.cleaned_data).first()
        if img_code.upper() != input_code.upper():
            form.add_error("code", "验证码错误")
            return render(request, 'login.html', {"form": form})
        # 不存在
        if not admin_object:
            form.add_error("password", "用户密码错误")
            return render(request, 'login.html', {"form": form})
        # 存在则登录成功，网站生成随机字符串，写到用户浏览器的cookie中，与服务器中的session进行校验
        request.session['info'] = {"id": admin_object.pk, "name": admin_object.name}
        return redirect("/admin/list")
    return render(request, 'login.html', {"form": form})


def sms_login(request):
    if request.method == "GET":
        form = SmsLoginForm()
        return render(request, "smslogin.html", {"form": form})
    form = SmsLoginForm(data=request.POST)
    if form.is_valid():
        mobile_phone = form.cleaned_data['phone']
        admin_object = Admin.objects.filter(phone=mobile_phone).first()
        print('登录成功')
        # 存在则登录成功，网站生成随机字符串，写到用户浏览器的cookie中，与服务器中的session进行校验
        request.session['info'] = {"id": admin_object.pk, "name": admin_object.name}
        return HttpResponse(json.dumps({"status": True, "data": '/admin/list/'}))
    return HttpResponse(json.dumps({"status": False, "error": form.errors}))


def logout(request):
    """ 注销 """
    # 清楚当前session
    request.session.clear()
    return redirect("/login/")


def register(request):
    """用户注册"""
    if request.method == "GET":
        form = RegisterForm()
        return render(request, "register.html", {"form": form})
    form = RegisterForm(data=request.POST)
    print(request.POST)
    if form.is_valid():
        # 存入数据库
        data_dic = form.cleaned_data
        data_dic.pop('code')
        data_dic.pop('confirm_password')
        Admin.objects.create(**data_dic)
        return HttpResponse(json.dumps({"status": True, "data": '/login/'}))
    return HttpResponse(json.dumps({"status": False, "error": form.errors}))


def send_sms(request):
    form = SendSmsForm(request, request.GET)
    if form.is_valid():
        return HttpResponse(json.dumps({"status": True}))
    print(form.errors)
    return HttpResponse(json.dumps({"status": False, "error": form.errors}))


def image_code(request):
    """ 生成图片验证码 """
    # 调用pillow函数,生成图片
    img, code_string = check_code()
    print('生成验证码:', code_string)
    # 写入到自己的session中,以便于后续获取验证码再进行校验
    request.session['image_code'] = code_string
    # 给session设置 60s 超时
    request.session.set_expiry(60 * 60 * 24)
    # 将图片保存到内存
    stream = BytesIO()
    img.save(stream, 'png')
    return HttpResponse(stream.getvalue())


def order_list(request):
    """
    订单管理列表
    :return:
    """
    form = OrderModelForm()
    data_dic = {}
    search_data = request.GET.get('query', '')
    if search_data:
        data_dic['title__contains'] = search_data
    queryset = Order.objects.filter(**data_dic).order_by('id')
    page_obj = Pagination(request, queryset, 10)
    # 分页后的数据
    page_data = page_obj.page_data
    # 分页html页面
    page_html = page_obj.html()
    context = {
        "queryset": page_data,
        "form": form,
        "page_string": page_html,
        "search_data": search_data,
    }
    return render(request, "order_list.html", context)


@csrf_exempt
def order_ajax(request):
    """
    ajax测试
    :param request:
    :return:
    """
    pwd = request.POST.get("pwd")
    user = request.POST.get("user")
    print(request.POST)
    if user == "root" and pwd == "root":
        data_dict = {"status": True}
    else:
        data_dict = {"status": False}
    return HttpResponse(json.dumps(data_dict))


@csrf_exempt
def order_add(request):
    """新建订单"""
    form = OrderModelForm(data=request.POST)
    print(request.POST)
    if form.is_valid():
        # 增加 oid 订单号
        form.instance.oid = datetime.datetime.now().strftime("%Y%m%d%H%M%S") + str(random.randint(1000, 9000))
        # 设置当前登录用户为订单的管理员
        admin_user = request.session["info"]["id"]
        form.instance.admin_id = admin_user
        form.save()
        return HttpResponse(json.dumps({"status": True}))
    return HttpResponse(json.dumps({"status": False, "error": form.errors}))


def order_delete(request):
    """ 删除订单 """
    uid = request.GET.get('uid')
    # 判断获取到的 id 数据行是否存在
    if Order.objects.filter(id=uid).exists():
        Order.objects.filter(id=uid).delete()
        return HttpResponse(json.dumps({"status": True}))
    else:
        return HttpResponse(json.dumps({"status": False, "error": "删除失败, 数据不存在,请刷新页面后重试!"}))


def order_detail(request):
    """ 根据id获取订单详情 """
    uid = request.GET.get('uid')
    row_object = Order.objects.filter(id=uid).values("title", "price", "status").first()
    print(uid, row_object)
    if not row_object:
        return HttpResponse(json.dumps({"status": False, "error": "数据不存在!"}))

    # 从数据库中获取到一个对象 row_object
    result = {
        "status": True,
        "data": row_object,
    }

    return HttpResponse(json.dumps(result))


@csrf_exempt
def order_edit(request):
    """ 编辑订单 """
    uid = request.GET.get('uid')
    row_object = Order.objects.filter(id=uid).first()
    print(uid, row_object)
    if not row_object:
        return HttpResponse(json.dumps({"status": False, "tips": "数据不存在!"}))

    # 获取编辑界面提交的数据
    data = request.POST
    # 判断编辑的内容是否跟以前一样
    if data["title"] != '' and data['price'] != '' and data['status'] != '':
        flag = 0
        if row_object.title == data["title"]:
            flag += 1
        if row_object.price == int(data["price"]):
            flag += 1
        if row_object.status == int(data["status"]):
            flag += 1
        # 如果输入的内容与原有内容相同,则进行提示
        if flag == 3:
            return HttpResponse(json.dumps({"status": False, "tips": "您没有变更任何内容,无需修改!"}))
    form = OrderModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return HttpResponse(json.dumps({"status": True}))

    return HttpResponse(json.dumps({"status": False, "error": form.errors}))
